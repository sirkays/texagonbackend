import csv
import io
import logging
import openpyxl
from decimal import Decimal, InvalidOperation

from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status
from rest_framework.decorators import (
    api_view,
    authentication_classes,
    permission_classes,
)
from rest_framework.response import Response
from rest_framework_api_key.permissions import HasAPIKey

from academics.models import StudentProfile, TeacherProfile, Classroom
from api.authentication import SessionTokenAuthentication
from core.utils import _resolve_org
from learning.models import Course, Enrollment
from orgs.models import AcademicSession

from .models import OfflinePracticalWork, OfflinePracticalScore
from .serializers import (
    OfflinePracticalWorkSerializer,
    OfflinePracticalScoreSerializer,
    OPWStudentSerializer,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_teacher(user):
    return getattr(user, "teacher_profile", None)


def _teacher_required(request):
    """Returns (teacher, None) or (None, Response)."""
    teacher = _get_teacher(request.user)
    if not teacher:
        return None, Response(
            {"detail": "Teacher profile not found."},
            status=status.HTTP_403_FORBIDDEN,
        )
    return teacher, None


def _get_opw_for_teacher(opw_id, teacher):
    """Fetch an OPW that belongs to the teacher's courses."""
    try:
        return OfflinePracticalWork.objects.select_related(
            "course", "academic_session"
        ).get(pk=opw_id, course__teacher=teacher)
    except OfflinePracticalWork.DoesNotExist:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# 1. LIST + CREATE  —  /opw/api/works/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def opw_list_create(request):
    teacher, err = _teacher_required(request)
    if err:
        return err

    org, _ = _resolve_org(request)

    if request.method == "GET":
        qs = OfflinePracticalWork.objects.filter(
            course__teacher=teacher
        ).select_related("course", "academic_session").order_by("-created_at")

        # Optional filter by course
        course_id = request.query_params.get("course_id")
        if course_id:
            qs = qs.filter(course_id=course_id)

        # Optional filter by assessment type
        atype = request.query_params.get("assessment_type")
        if atype:
            qs = qs.filter(assessment_type=atype)

        serializer = OfflinePracticalWorkSerializer(qs, many=True)
        return Response(serializer.data)

    # POST — create
    data = request.data.copy()
    serializer = OfflinePracticalWorkSerializer(data=data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Attach current session automatically
    current_session = None
    if org:
        current_session = AcademicSession.objects.filter(
            organization=org, is_current=True
        ).first()

    opw = serializer.save(
        created_by=teacher,
        academic_session=current_session,
    )
    return Response(
        OfflinePracticalWorkSerializer(opw).data,
        status=status.HTTP_201_CREATED,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 2. RETRIEVE + UPDATE + DELETE  —  /opw/api/works/<id>/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def opw_detail(request, opw_id):
    teacher, err = _teacher_required(request)
    if err:
        return err

    opw = _get_opw_for_teacher(opw_id, teacher)
    if not opw:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(OfflinePracticalWorkSerializer(opw).data)

    if request.method == "PATCH":
        serializer = OfflinePracticalWorkSerializer(opw, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data)

    if request.method == "DELETE":
        opw.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ─────────────────────────────────────────────────────────────────────────────
# 3. STUDENT LIST WITH SCORES  —  /opw/api/works/<id>/students/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def opw_students(request, opw_id):
    """
    Returns enrolled students for the OPW's course together with their
    current score (if any).

    Query params:
      - classroom (int): filter by classroom id
      - search (str): filter by student name/email
      - page (int): page number, default 1
      - page_size (int): page size, default 20
    """
    teacher, err = _teacher_required(request)
    if err:
        return err

    opw = _get_opw_for_teacher(opw_id, teacher)
    if not opw:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    # All actively enrolled students for this course
    enrolled_ids = Enrollment.objects.filter(
        course=opw.course,
        status=Enrollment.Status.ACTIVE,
    ).values_list("student_id", flat=True)

    qs = StudentProfile.objects.filter(
        pk__in=enrolled_ids
    ).select_related("user", "current_classroom").order_by(
        "user__first_name", "user__last_name"
    )

    # Filter by classroom
    classroom_id = request.query_params.get("classroom")
    if classroom_id:
        qs = qs.filter(current_classroom_id=classroom_id)

    # Search by name or email
    search = request.query_params.get("search", "").strip()
    if search:
        qs = qs.filter(
            Q(user__first_name__icontains=search)
            | Q(user__last_name__icontains=search)
            | Q(user__email__icontains=search)
        )

    # Paginate
    page = int(request.query_params.get("page", 1))
    page_size = int(request.query_params.get("page_size", 20))
    paginator = Paginator(qs, page_size)
    page_obj = paginator.get_page(page)

    # Fetch existing scores for the current page in one query
    student_ids = [s.id for s in page_obj.object_list]
    scores_map = {
        s.student_id: s
        for s in OfflinePracticalScore.objects.filter(
            opw=opw, student_id__in=student_ids
        )
    }

    results = []
    for student in page_obj.object_list:
        user = student.user
        score_obj = scores_map.get(student.id)
        results.append({
            "student_id": student.id,
            "student_name": f"{user.first_name} {user.last_name}".strip() or user.email,
            "student_email": user.email,
            "classroom_id": student.current_classroom_id,
            "classroom_name": student.current_classroom.name if student.current_classroom else None,
            "score_id": score_obj.id if score_obj else None,
            "score": str(score_obj.score) if score_obj and score_obj.score is not None else None,
            "feedback": score_obj.feedback if score_obj else None,
            "recorded_at": score_obj.recorded_at.isoformat() if score_obj else None,
        })

    return Response({
        "results": results,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": paginator.count,
            "pages": paginator.num_pages,
        },
    })


# ─────────────────────────────────────────────────────────────────────────────
# 4. SUBMIT / UPSERT SCORES  —  /opw/api/works/<id>/scores/submit/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
@transaction.atomic
def opw_submit_scores(request, opw_id):
    """
    Upsert one or more student scores.
    Body:
      { "scores": [ { "student_id": 5, "score": 87.5, "feedback": "..." }, ... ] }

    Score can be null to "un-score" a student.
    """
    teacher, err = _teacher_required(request)
    if err:
        return err

    opw = _get_opw_for_teacher(opw_id, teacher)
    if not opw:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    scores_data = request.data.get("scores", [])
    if not isinstance(scores_data, list) or len(scores_data) == 0:
        return Response(
            {"detail": "scores must be a non-empty list."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    saved = []
    errors = []

    for item in scores_data:
        student_id = item.get("student_id")
        raw_score = item.get("score")
        feedback = item.get("feedback", "")

        if not student_id:
            errors.append({"student_id": student_id, "error": "student_id is required."})
            continue

        # Validate score range
        score_val = None
        if raw_score is not None and str(raw_score).strip() != "":
            try:
                score_val = Decimal(str(raw_score))
                if score_val < 0 or score_val > opw.max_score:
                    errors.append({
                        "student_id": student_id,
                        "error": f"Score must be between 0 and {opw.max_score}.",
                    })
                    continue
            except InvalidOperation:
                errors.append({"student_id": student_id, "error": "Invalid score value."})
                continue

        # Upsert
        obj, created = OfflinePracticalScore.objects.update_or_create(
            opw=opw,
            student_id=student_id,
            defaults={
                "score": score_val,
                "feedback": feedback,
                "recorded_at": timezone.now(),
                "recorded_by": teacher,
            },
        )
        saved.append({
            "student_id": student_id,
            "score_id": obj.id,
            "score": str(obj.score) if obj.score is not None else None,
            "created": created,
        })

    return Response(
        {"saved": saved, "errors": errors},
        status=status.HTTP_200_OK,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 5. EXPORT SCORES AS CSV  —  /opw/api/works/<id>/scores/export/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def opw_export_csv(request, opw_id):
    teacher, err = _teacher_required(request)
    if err:
        return err

    opw = _get_opw_for_teacher(opw_id, teacher)
    if not opw:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    scores = OfflinePracticalScore.objects.filter(opw=opw).select_related(
        "student__user", "student__classroom"
    ).order_by("student__user__first_name", "student__user__last_name")

    output = io.StringIO()
    writer = csv.writer(output)

    # Header section
    writer.writerow(["OFF-PRACTICAL WORK — SCORE EXPORT"])
    writer.writerow(["Title", opw.title])
    writer.writerow(["Type", opw.get_assessment_type_display()])
    writer.writerow(["Course", opw.course.name])
    writer.writerow(["Max Score", str(opw.max_score)])
    writer.writerow(["Conducted", str(opw.conducted_at) if opw.conducted_at else ""])
    writer.writerow(["Exported", timezone.now().strftime("%Y-%m-%d %H:%M")])
    writer.writerow([])

    # Data columns
    writer.writerow(["#", "Student Name", "Email", "Classroom", "Score", "Max Score", "Percentage", "Grade", "Feedback", "Recorded At"])

    for idx, s in enumerate(scores, 1):
        user = s.student.user
        name = f"{user.first_name} {user.last_name}".strip() or user.email
        classroom = s.student.classroom.name if s.student.classroom else ""
        score_val = float(s.score) if s.score is not None else None
        max_val = float(opw.max_score)
        pct = round(score_val / max_val * 100, 1) if score_val is not None else ""
        grade = ""
        if score_val is not None:
            p = pct
            grade = "A" if p >= 90 else "B" if p >= 80 else "C" if p >= 70 else "D" if p >= 50 else "F"

        writer.writerow([
            idx,
            name,
            user.email,
            classroom,
            score_val if score_val is not None else "Not Scored",
            max_val,
            f"{pct}%" if pct != "" else "",
            grade,
            s.feedback,
            s.recorded_at.strftime("%Y-%m-%d %H:%M") if s.recorded_at else "",
        ])

    safe_title = opw.title.replace(" ", "_")[:40]
    filename = f"OPW_{safe_title}_{timezone.now().strftime('%Y%m%d')}.csv"

    response = HttpResponse(
        "\ufeff" + output.getvalue(),  # BOM for Excel compatibility
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# 6. TEACHER COURSES  —  /opw/api/courses/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def opw_teacher_courses(request):
    teacher, err = _teacher_required(request)
    if err:
        return err

    courses = Course.objects.filter(teacher=teacher).order_by("name")
    data = [{"id": c.id, "title": c.name, "course_type": c.course_type} for c in courses]
    return Response(data)


# ─────────────────────────────────────────────────────────────────────────────
# 7. CLASSROOMS FOR A COURSE  —  /opw/api/classrooms/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def opw_course_classrooms(request):
    """
    Returns classrooms that have at least one enrolled student in the given course.
    Query param: course_id (required)
    """
    teacher, err = _teacher_required(request)
    if err:
        return err

    course_id = request.query_params.get("course_id")
    if not course_id:
        return Response({"detail": "course_id is required."}, status=status.HTTP_400_BAD_REQUEST)

    # Verify teacher owns this course
    try:
        course = Course.objects.get(pk=course_id, teacher=teacher)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

    # Find classrooms from enrolled students
    classroom_ids = StudentProfile.objects.filter(
        pk__in=Enrollment.objects.filter(
            course=course,
            status=Enrollment.Status.ACTIVE,
        ).values_list("student_id", flat=True),
        current_classroom__isnull=False,
    ).values_list("current_classroom_id", flat=True).distinct()

    classrooms = Classroom.objects.filter(pk__in=classroom_ids).order_by("name")
    data = [{"id": c.id, "name": c.name} for c in classrooms]
    return Response(data)

# ─────────────────────────────────────────────────────────────────────────────
# 8. EXPORT SCORES AS EXCEL  —  /opw/api/works/<id>/scores/export-excel/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def opw_export_excel(request, opw_id):
    teacher, err = _teacher_required(request)
    if err:
        return err

    opw = _get_opw_for_teacher(opw_id, teacher)
    if not opw:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    # All actively enrolled students for this course
    enrolled_ids = Enrollment.objects.filter(
        course=opw.course,
        status=Enrollment.Status.ACTIVE,
    ).values_list("student_id", flat=True)

    students = StudentProfile.objects.filter(
        pk__in=enrolled_ids
    ).select_related("user", "current_classroom").order_by(
        "user__first_name", "user__last_name"
    )

    scores_map = {
        s.student_id: s
        for s in OfflinePracticalScore.objects.filter(opw=opw, student_id__in=enrolled_ids)
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OPW Scores"

    headers = ["Student ID", "Student Name", "Email", "Classroom", "Score", "Feedback"]
    ws.append(headers)

    for student in students:
        user = student.user
        name = f"{user.first_name} {user.last_name}".strip() or user.email
        classroom = student.current_classroom.name if student.current_classroom else ""
        score_obj = scores_map.get(student.id)
        
        score_val = float(score_obj.score) if (score_obj and score_obj.score is not None) else ""
        feedback = score_obj.feedback if score_obj else ""
        
        ws.append([
            student.id,
            name,
            user.email,
            classroom,
            score_val,
            feedback
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = opw.title.replace(" ", "_")[:40]
    filename = f"OPW_{safe_title}_{timezone.now().strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# 9. IMPORT SCORES FROM EXCEL  —  /opw/api/works/<id>/scores/import-excel/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def opw_import_excel(request, opw_id):
    teacher, err = _teacher_required(request)
    if err:
        return err

    opw = _get_opw_for_teacher(opw_id, teacher)
    if not opw:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        
    if not file_obj.name.endswith(".xlsx"):
        return Response({"error": "Please upload a valid .xlsx file."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Excel read error: {e}")
        return Response({"error": "Failed to read Excel file. It might be corrupted."}, status=status.HTTP_400_BAD_REQUEST)

    # Find headers
    headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value is not None]
    
    try:
        id_idx = headers.index("student id")
    except ValueError:
        return Response({"error": "'Student ID' column not found in row 1."}, status=status.HTTP_400_BAD_REQUEST)

    score_idx = headers.index("score") if "score" in headers else None
    feedback_idx = headers.index("feedback") if "feedback" in headers else None

    if score_idx is None:
        return Response({"error": "'Score' column not found in row 1."}, status=status.HTTP_400_BAD_REQUEST)

    saved_count = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        student_id_val = row[id_idx].value
        if not student_id_val:
            continue
            
        try:
            student_id = int(student_id_val)
        except ValueError:
            errors.append({"row": row_idx, "error": f"Invalid Student ID: {student_id_val}"})
            continue
            
        score_val_raw = row[score_idx].value
        feedback_val = row[feedback_idx].value if feedback_idx is not None else ""
        
        if score_val_raw is None or str(score_val_raw).strip() == "":
            continue # Skip blank scores
            
        try:
            score_val = Decimal(str(score_val_raw))
        except InvalidOperation:
            errors.append({"row": row_idx, "error": f"Invalid Score: {score_val_raw} for Student {student_id}"})
            continue

        if score_val < 0 or score_val > opw.max_score:
            errors.append({"row": row_idx, "error": f"Score {score_val} is outside valid range (0 - {opw.max_score}) for Student {student_id}"})
            continue
            
        OfflinePracticalScore.objects.update_or_create(
            opw=opw,
            student_id=student_id,
            defaults={
                "score": score_val,
                "feedback": str(feedback_val) if feedback_val is not None else "",
                "recorded_at": timezone.now(),
                "recorded_by": teacher,
            },
        )
        saved_count += 1

    return Response({
        "message": f"Successfully updated {saved_count} scores.",
        "saved_count": saved_count,
        "errors": errors
    }, status=status.HTTP_200_OK)
