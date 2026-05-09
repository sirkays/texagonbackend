# academics/management/commands/attach_students_to_classrooms.py

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Optional, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from openpyxl import load_workbook

from accounts.models import User
from academics.models import Classroom, StudentProfile
from orgs.models import Organization, OrganizationMembership


class Command(BaseCommand):
    help = (
        "Attach existing student profiles to their respective classrooms from an Excel file. "
        "Creates classrooms for the organization if they do not already exist."
    )

    REQUIRED_COLUMNS = {
        "email": ["email", "student email", "email address"],
        "class": ["class", "classroom", "student class", "class name"],
    }

    OPTIONAL_COLUMNS = {
        "student_name": ["student name", "name", "full name"],
        "admission_no": ["admission no", "admission number", "admission_no"],
    }

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Path to the Excel file. Example: /path/to/student_email_class_list.xlsx",
        )

        parser.add_argument(
            "--org-slug",
            required=True,
            help="Organization slug. Example: acss-kubwa",
        )

        parser.add_argument(
            "--sheet",
            default=None,
            help="Optional sheet name. If omitted, the first worksheet will be used.",
        )

        parser.add_argument(
            "--class-type",
            default="public",
            choices=["public", "private"],
            help="Classroom type to use when creating new classrooms. Default: public",
        )

        parser.add_argument(
            "--create-missing-profiles",
            action="store_true",
            help="Create StudentProfile if the user exists but student profile is missing.",
        )

        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Preview changes without saving anything.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        org_slug = options["org_slug"]
        sheet_name = options["sheet"]
        class_type = options["class_type"]
        create_missing_profiles = options["create_missing_profiles"]
        dry_run = options["dry_run"]

        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        try:
            organization = Organization.objects.get(slug=org_slug)
        except Organization.DoesNotExist:
            raise CommandError(f'Organization with slug "{org_slug}" does not exist.')

        workbook = load_workbook(filename=file_path, data_only=True)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(
                    f'Sheet "{sheet_name}" not found. Available sheets: {", ".join(workbook.sheetnames)}'
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        header_map = self.get_header_map(worksheet)

        email_col = self.find_column(header_map, self.REQUIRED_COLUMNS["email"])
        class_col = self.find_column(header_map, self.REQUIRED_COLUMNS["class"])

        if email_col is None:
            raise CommandError(
                "Could not find an Email column. Expected one of: "
                + ", ".join(self.REQUIRED_COLUMNS["email"])
            )

        if class_col is None:
            raise CommandError(
                "Could not find a Class column. Expected one of: "
                + ", ".join(self.REQUIRED_COLUMNS["class"])
            )

        name_col = self.find_column(header_map, self.OPTIONAL_COLUMNS["student_name"])
        admission_col = self.find_column(header_map, self.OPTIONAL_COLUMNS["admission_no"])

        stats = {
            "rows_read": 0,
            "attached": 0,
            "classrooms_created": 0,
            "classrooms_existing": 0,
            "profiles_created": 0,
            "users_missing": 0,
            "profiles_missing": 0,
            "skipped_no_email": 0,
            "skipped_no_class": 0,
            "errors": 0,
        }

        classroom_cache: Dict[str, Classroom] = {}
        classroom_created_cache: Dict[str, bool] = {}

        self.stdout.write("")
        self.stdout.write(self.style.WARNING("Starting classroom attachment import..."))

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN ENABLED: no database changes will be saved."))

        try:
            with transaction.atomic():
                for row_number in range(2, worksheet.max_row + 1):
                    stats["rows_read"] += 1

                    email = self.clean_email(worksheet.cell(row=row_number, column=email_col).value)
                    class_name = self.clean_text(worksheet.cell(row=row_number, column=class_col).value)

                    student_name = (
                        self.clean_text(worksheet.cell(row=row_number, column=name_col).value)
                        if name_col
                        else ""
                    )

                    admission_no = (
                        self.clean_text(worksheet.cell(row=row_number, column=admission_col).value)
                        if admission_col
                        else ""
                    )

                    if not email:
                        stats["skipped_no_email"] += 1
                        self.stdout.write(
                            self.style.WARNING(f"Row {row_number}: skipped because email is empty.")
                        )
                        continue

                    if not class_name:
                        stats["skipped_no_class"] += 1
                        self.stdout.write(
                            self.style.WARNING(f"Row {row_number}: skipped {email} because class is empty.")
                        )
                        continue

                    try:
                        user = User.objects.get(email__iexact=email)
                    except User.DoesNotExist:
                        stats["users_missing"] += 1
                        self.stdout.write(
                            self.style.ERROR(
                                f"Row {row_number}: user not found for email {email}. "
                                f"Student: {student_name or 'N/A'}"
                            )
                        )
                        continue
                    except User.MultipleObjectsReturned:
                        stats["errors"] += 1
                        self.stdout.write(
                            self.style.ERROR(f"Row {row_number}: multiple users found for email {email}.")
                        )
                        continue

                    student_profile = getattr(user, "student_profile", None)

                    if not student_profile:
                        if not create_missing_profiles:
                            stats["profiles_missing"] += 1
                            self.stdout.write(
                                self.style.ERROR(
                                    f"Row {row_number}: StudentProfile missing for {email}. "
                                    f"Use --create-missing-profiles if you want to create it."
                                )
                            )
                            continue

                        student_profile = StudentProfile(
                            user=user,
                            organization=organization,
                            admission_no=admission_no or None,
                        )

                        if not dry_run:
                            student_profile.save()

                        stats["profiles_created"] += 1

                    classroom_key = self.normalize_class_key(class_name)

                    if classroom_key in classroom_cache:
                        classroom = classroom_cache[classroom_key]
                        created = classroom_created_cache[classroom_key]
                    else:
                        classroom, created = self.get_or_create_classroom(
                            organization=organization,
                            class_name=class_name,
                            class_type=class_type,
                            dry_run=dry_run,
                        )

                        classroom_cache[classroom_key] = classroom
                        classroom_created_cache[classroom_key] = created

                        if created:
                            stats["classrooms_created"] += 1
                        else:
                            stats["classrooms_existing"] += 1

                    if not dry_run:
                        update_fields = []

                        if student_profile.organization_id != organization.id:
                            student_profile.organization = organization
                            update_fields.append("organization")

                        if student_profile.current_classroom_id != classroom.id:
                            student_profile.current_classroom = classroom
                            update_fields.append("current_classroom")

                        if admission_no and not student_profile.admission_no:
                            student_profile.admission_no = admission_no
                            update_fields.append("admission_no")

                        if update_fields:
                            student_profile.save(update_fields=update_fields)

                        user_update_fields = []

                        if user.primary_org_id != organization.id:
                            user.primary_org = organization
                            user_update_fields.append("primary_org")

                        if user_update_fields:
                            user.save(update_fields=user_update_fields)

                        OrganizationMembership.objects.get_or_create(
                            user=user,
                            organization=organization,
                            role=OrganizationMembership.Role.STUDENT,
                            defaults={"is_active": True},
                        )

                    stats["attached"] += 1

                    self.stdout.write(
                        self.style.SUCCESS(
                            f"Row {row_number}: attached {email} to {classroom.name}"
                        )
                    )

                if dry_run:
                    raise RollbackDryRun()

        except RollbackDryRun:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Dry run completed. All changes rolled back."))

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Import completed."))
        self.stdout.write("-" * 50)

        for key, value in stats.items():
            self.stdout.write(f"{key}: {value}")

    def get_header_map(self, worksheet) -> Dict[str, int]:
        """
        Reads first row and returns normalized header -> column index.
        """
        header_map = {}

        for cell in worksheet[1]:
            value = self.normalize_header(cell.value)
            if value:
                header_map[value] = cell.column

        return header_map

    def find_column(self, header_map: Dict[str, int], possible_names: list[str]) -> Optional[int]:
        for name in possible_names:
            normalized = self.normalize_header(name)
            if normalized in header_map:
                return header_map[normalized]
        return None

    def get_or_create_classroom(
        self,
        *,
        organization: Organization,
        class_name: str,
        class_type: str,
        dry_run: bool,
    ) -> Tuple[Classroom, bool]:
        """
        Find classroom by organization and case-insensitive name.
        If it does not exist, create it.

        Classroom unique_together is organization + name + code,
        so this avoids creating duplicate class names for the same org.
        """
        existing = (
            Classroom.objects.filter(
                organization=organization,
                name__iexact=class_name,
            )
            .order_by("id")
            .first()
        )

        if existing:
            return existing, False

        code = self.make_classroom_code(class_name)

        if dry_run:
            classroom = Classroom(
                id=-abs(hash((organization.id, class_name))) % 1000000,
                organization=organization,
                name=class_name,
                code=code,
                class_type=class_type,
            )
            return classroom, True

        classroom = Classroom.objects.create(
            organization=organization,
            name=class_name,
            code=code,
            class_type=class_type,
        )

        return classroom, True

    def make_classroom_code(self, class_name: str) -> str:
        """
        Converts class names like 'JSS 1 Gold' to 'JSS-1-GOLD'.
        """
        code = slugify(class_name).upper()
        code = code.replace("-", "_")
        return code[:32] or "CLASS"

    def clean_email(self, value) -> str:
        if value is None:
            return ""

        email = str(value).strip().lower()
        email = re.sub(r"\s+", "", email)

        return email

    def clean_text(self, value) -> str:
        if value is None:
            return ""

        text = str(value).strip()
        text = re.sub(r"\s+", " ", text)

        return text

    def normalize_header(self, value) -> str:
        if value is None:
            return ""

        text = str(value).strip().lower()
        text = text.replace("_", " ")
        text = re.sub(r"\s+", " ", text)

        return text

    def normalize_class_key(self, value) -> str:
        value = self.clean_text(value).lower()
        value = re.sub(r"[^a-z0-9]+", "", value)
        return value


class RollbackDryRun(Exception):
    pass

