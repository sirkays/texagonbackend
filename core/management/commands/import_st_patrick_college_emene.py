"""
Django management command: import_st_patrick_college_emene

SETUP — do this once before running the command:
=========================================================
1. Pick any existing Django app in your project (e.g. "academics", "core", "accounts").
2. Inside that app create this exact folder layout (each folder needs its own __init__.py):

   your_app/
   ├── management/
   │   ├── __init__.py          <- empty file, REQUIRED
   │   └── commands/
   │       ├── __init__.py      <- empty file, REQUIRED
   │       └── import_st_patrick_college_emene.py   <- THIS FILE

3. Make sure 'your_app' is listed in INSTALLED_APPS in settings.py.

Quick shell setup (run from your project root):
   mkdir -p your_app/management/commands
   touch your_app/management/__init__.py
   touch your_app/management/commands/__init__.py
   cp import_st_patrick_college_emene.py your_app/management/commands/

Then verify Django can see the command:
   python manage.py help | grep import_st_patrick
   (nothing shown = __init__.py missing OR app not in INSTALLED_APPS)
=========================================================

Usage:
  python manage.py import_st_patrick_college_emene /path/to/file.xlsx --org-id <PK>

Optional flags:
  --org-id    Organization PK  (required)
  --plan-id   SubscriptionPlan PK (default: 2)
  --dry-run   Parse without writing to DB
"""

import re
import unicodedata
from datetime import timedelta

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import User
from academics.models import Classroom, StudentProfile
from orgs.models import Organization, OrganizationMembership
from billing.models import SubscriptionPlan, UserAccountSubscription


def sanitise_for_email(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_only = nfkd.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", ascii_only.lower())


class Command(BaseCommand):
    help = "Import students from the St Patrick College Emene Excel file"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to the .xlsx file")
        parser.add_argument("--org-id",  type=int, required=True, help="Organization PK")
        parser.add_argument("--plan-id", type=int, default=2,     help="SubscriptionPlan PK (default: 2)")
        parser.add_argument("--dry-run", action="store_true",     help="Parse without writing to DB")

    def handle(self, *args, **options):
        excel_path = options["excel_path"]
        org_id     = options["org_id"]
        plan_id    = options["plan_id"]
        dry_run    = options["dry_run"]

        try:
            org = Organization.objects.get(pk=org_id)
        except Organization.DoesNotExist:
            raise CommandError(f"Organization pk={org_id} does not exist.")

        try:
            plan = SubscriptionPlan.objects.get(pk=plan_id)
        except SubscriptionPlan.DoesNotExist:
            raise CommandError(f"SubscriptionPlan pk={plan_id} does not exist.")

        try:
            wb = load_workbook(excel_path, read_only=True)
        except Exception as exc:
            raise CommandError(f"Cannot open Excel file: {exc}")

        password = "Techxagon@2026"
        now      = timezone.now()
        end_at   = now + timedelta(days=90)

        # Parse: row 0 = header, row 1+ = names in col A, tab name = classroom
        sheets_data = []
        for sheet_name in wb.sheetnames:
            ws             = wb[sheet_name]
            rows           = list(ws.iter_rows(values_only=True))
            classroom_name = sheet_name.strip()
            students       = []
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                if name:
                    students.append(name)
            if students:
                sheets_data.append({"classroom_name": classroom_name, "students": students})
        wb.close()

        total_parsed = sum(len(s["students"]) for s in sheets_data)
        self.stdout.write(
            f"\nParsed {total_parsed} students across {len(sheets_data)} sheets "
            f"(org_id={org_id}, plan_id={plan_id})"
        )

        if dry_run:
            self.stdout.write(self.style.WARNING("\n--dry-run: nothing will be written.\n"))
            for sd in sheets_data:
                self.stdout.write(f"  {sd['classroom_name']}: {len(sd['students'])} students")
            return

        # Pre-load existing emails to detect collisions without extra queries
        existing_emails: set = set(
            User.objects.filter(email__endswith="@techxagonacademy.com")
            .values_list("email", flat=True)
        )
        assigned_emails: set = set()

        created_users = created_profiles = created_memberships = 0
        created_subs  = created_classrooms = skipped_existing  = 0
        classroom_cache: dict = {}

        with transaction.atomic():
            for sd in sheets_data:
                classroom_name = sd["classroom_name"]

                if classroom_name not in classroom_cache:
                    classroom, cls_created = Classroom.objects.get_or_create(
                        organization=org,
                        name=classroom_name,
                        defaults={"code": ""},
                    )
                    if cls_created:
                        created_classrooms += 1
                    classroom_cache[classroom_name] = classroom
                else:
                    classroom = classroom_cache[classroom_name]

                sheet_created = 0

                for full_name in sd["students"]:

                    base_local = sanitise_for_email(full_name) or "student"
                    local_part = base_local + "spc"
                    candidate  = f"{local_part}@techxagonacademy.com"
                    suffix_num = 0
                    while candidate in existing_emails or candidate in assigned_emails:
                        suffix_num += 1
                        candidate   = f"{local_part}{suffix_num}@techxagonacademy.com"

                    email = candidate
                    assigned_emails.add(email)
                    existing_emails.add(email)

                    if User.objects.filter(email=email).exists():
                        skipped_existing += 1
                        continue

                    parts      = full_name.title().split()
                    first_name = parts[0] if parts else ""
                    last_name  = " ".join(parts[1:]) if len(parts) > 1 else ""

                    user = User.objects.create_user(
                        email=email,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                        is_generated=True,
                        is_active=True,
                        primary_org=org,
                    )
                    created_users += 1

                    OrganizationMembership.objects.create(
                        user=user,
                        organization=org,
                        role=OrganizationMembership.Role.STUDENT,
                        is_active=True,
                    )
                    created_memberships += 1

                    StudentProfile.objects.create(
                        user=user,
                        organization=org,
                        current_classroom=classroom,
                    )
                    created_profiles += 1

                    UserAccountSubscription.objects.create(
                        organization=org,
                        user=user,
                        plan=plan,
                        status=UserAccountSubscription.Status.ACTIVE,
                        start_at=now,
                        end_at=end_at,
                        amount=plan.price,
                        currency="NGN",
                    )
                    created_subs  += 1
                    sheet_created += 1

                self.stdout.write(
                    f"  ✔ {classroom_name}: {sheet_created}/{len(sd['students'])} created"
                )

        self.stdout.write(self.style.SUCCESS(
            f"\n{'='*52}\n"
            f"  Classrooms created  : {created_classrooms}\n"
            f"  Users created       : {created_users}\n"
            f"  Memberships created : {created_memberships}\n"
            f"  Profiles created    : {created_profiles}\n"
            f"  Subscriptions       : {created_subs}\n"
            f"  Skipped (existing)  : {skipped_existing}\n"
            f"{'='*52}"
        ))