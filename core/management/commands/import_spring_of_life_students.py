"""
Django management command: import_spring_of_life_students

Place this file at:
  <your_app>/management/commands/import_spring_of_life_students.py

Make sure the management/ and management/commands/ directories each
contain an __init__.py file.

Usage:
  python manage.py import_spring_of_life_students /path/to/student_list.xlsx

What it does (all inside a single atomic transaction):
  1. Reads every sheet from the Excel file.
     - Skips duplicate classrooms (detected via the header-row classroom name).
  2. For each sheet → creates a Classroom (if it doesn't already exist)
     under Organization pk=8.
  3. For each student row:
     a. Creates a User (email = <sanitised-full-name>sop@techxagonacademy.com,
        password = Techxagon@2026, is_generated=True).
        If the email already exists it appends a numeric suffix.
     b. Creates an OrganizationMembership (role=student, org=8).
     c. Creates a StudentProfile linked to the classroom.
     d. Creates a UserAccountSubscription (plan pk=2, 3 months, active).
  4. Prints a per-sheet summary and a grand total.
"""

import re
import unicodedata
from datetime import timedelta

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import User
from academics.models import Classroom, StudentProfile
from orgs.models import Organization, OrganizationMembership
from billing.models import SubscriptionPlan, UserAccountSubscription


def sanitise_for_email(full_name: str) -> str:
    """
    Turn a full name into a lowercase email-safe local part.
    - NFKD-normalise (strips accents / special Unicode like Greek chars)
    - Keep only [a-z0-9]
    - Example: "Agu Dolly Chimamanda" → "agudollychimamanda"
    """
    nfkd = unicodedata.normalize("NFKD", full_name)
    ascii_only = nfkd.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", ascii_only.lower())


class Command(BaseCommand):
    help = "Import students from the Spring-of-Life Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Path to the .xlsx student list file",
        )
        parser.add_argument(
            "--org-id",
            type=int,
            default=8,
            help="Organization PK (default: 8)",
        )
        parser.add_argument(
            "--plan-id",
            type=int,
            default=2,
            help="SubscriptionPlan PK (default: 2)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database",
        )

    def handle(self, *args, **options):
        excel_path = options["excel_path"]
        org_id = options["org_id"]
        plan_id = options["plan_id"]
        dry_run = options["dry_run"]

        # ── Pre-flight checks ──────────────────────────────────────
        try:
            org = Organization.objects.get(pk=org_id)
        except Organization.DoesNotExist:
            raise CommandError(f"Organization with pk={org_id} does not exist.")

        try:
            plan = SubscriptionPlan.objects.get(pk=plan_id)
        except SubscriptionPlan.DoesNotExist:
            raise CommandError(f"SubscriptionPlan with pk={plan_id} does not exist.")

        try:
            wb = load_workbook(excel_path, read_only=True)
        except Exception as exc:
            raise CommandError(f"Cannot open Excel file: {exc}")

        password = "Techxagon@2026"
        now = timezone.now()
        end_at = now + timedelta(days=90)  # 3 months

        # ── Collect all emails we will generate so we can detect
        #    duplicates both within the file AND against the DB. ─────
        existing_emails = set(
            User.objects.filter(email__endswith="@techxagonacademy.com")
            .values_list("email", flat=True)
        )
        # Track emails assigned during THIS run (for intra-file dupes)
        assigned_emails: dict[str, str] = {}  # email → full_name (for logging)

        # ── Parse sheets ───────────────────────────────────────────
        seen_classroom_names: set[str] = set()
        sheets_data: list[dict] = []  # [{classroom_name, sheet_tab, students: [{name, gender}]}]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows or not rows[0] or len(rows[0]) < 1:
                continue

            header_cell = str(rows[0][0] or "")
            classroom_name = header_cell.replace("Name", "").strip()
            if not classroom_name:
                self.stderr.write(f"  ⚠ Sheet '{sheet_name}': empty header – skipped")
                continue

            if classroom_name in seen_classroom_names:
                self.stdout.write(f"  ⏭ Sheet '{sheet_name}' duplicate of '{classroom_name}' – skipped")
                continue
            seen_classroom_names.add(classroom_name)

            students = []
            for row in rows[1:]:
                if not row or len(row) < 1:
                    continue
                name = str(row[0] or "").strip()
                if not name:
                    continue
                gender = str(row[1] or "").strip() if len(row) > 1 else ""
                students.append({"name": name, "gender": gender})

            sheets_data.append({
                "classroom_name": classroom_name,
                "sheet_tab": sheet_name,
                "students": students,
            })

        wb.close()

        total_parsed = sum(len(s["students"]) for s in sheets_data)
        self.stdout.write(
            f"\nParsed {total_parsed} students across {len(sheets_data)} sheets "
            f"(org={org_id}, plan={plan_id})"
        )

        if dry_run:
            self.stdout.write(self.style.WARNING("\n--dry-run enabled. Nothing written.\n"))
            for sd in sheets_data:
                self.stdout.write(f"  {sd['classroom_name']}: {len(sd['students'])} students")
            return

        # ── Write to DB inside one transaction ─────────────────────
        created_users = 0
        created_profiles = 0
        created_memberships = 0
        created_subs = 0
        created_classrooms = 0
        skipped_existing = 0

        with transaction.atomic():
            for sd in sheets_data:
                classroom_name = sd["classroom_name"]

                # Get or create Classroom
                classroom, cr = Classroom.objects.get_or_create(
                    organization=org,
                    name=classroom_name,
                    defaults={"code": ""},
                )
                if cr:
                    created_classrooms += 1

                sheet_created = 0

                for stu in sd["students"]:
                    full_name = stu["name"]

                    # ── Build unique email ──────────────────────
                    base_local = sanitise_for_email(full_name)
                    if not base_local:
                        base_local = "student"
                    local_part = base_local + "sop"
                    email = f"{local_part}@techxagonacademy.com"

                    # Deduplicate against DB + this run
                    suffix = 0
                    candidate = email
                    while candidate in existing_emails or candidate in assigned_emails:
                        suffix += 1
                        candidate = f"{local_part}{suffix}@techxagonacademy.com"
                    email = candidate
                    assigned_emails[email] = full_name
                    existing_emails.add(email)

                    # ── Split name into first / last ────────────
                    parts = full_name.strip().split()
                    first_name = parts[0] if parts else ""
                    last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

                    # ── Check if user already exists by email ───
                    if User.objects.filter(email=email).exists():
                        skipped_existing += 1
                        continue

                    # ── Create User (username auto-generated by UserManager) ─
                    user = User.objects.create_user(
                        email=email,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                        is_generated=True,
                        is_active=True,
                    )
                    created_users += 1

                    # ── OrganizationMembership ──────────────────
                    OrganizationMembership.objects.create(
                        user=user,
                        organization=org,
                        role=OrganizationMembership.Role.STUDENT,
                        is_active=True,
                    )
                    created_memberships += 1

                    # ── StudentProfile ──────────────────────────
                    StudentProfile.objects.create(
                        user=user,
                        organization=org,
                        current_classroom=classroom,
                    )
                    created_profiles += 1

                    # ── UserAccountSubscription (3 months) ──────
                    UserAccountSubscription.objects.create(
                        organization=org,
                        user=user,
                        plan=plan,
                        status=UserAccountSubscription.Status.ACTIVE,
                        start_at=now,
                        end_at=end_at,
                        amount=plan.price,
                        currency="NGN",
                    )
                    created_subs += 1
                    sheet_created += 1

                self.stdout.write(
                    f"  ✔ {classroom_name}: {sheet_created}/{len(sd['students'])} created"
                )

        # ── Summary ────────────────────────────────────────────────
        self.stdout.write(self.style.SUCCESS(
            f"\n{'='*50}\n"
            f"  Classrooms created : {created_classrooms}\n"
            f"  Users created      : {created_users}\n"
            f"  Memberships created: {created_memberships}\n"
            f"  Profiles created   : {created_profiles}\n"
            f"  Subscriptions      : {created_subs}\n"
            f"  Skipped (existing) : {skipped_existing}\n"
            f"{'='*50}"
        ))