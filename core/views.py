# views.py
from typing import Optional, Dict, Any, List
import traceback

from django.conf import settings
from django.db import transaction
from django.db.models import Q, Count, Max
from django.shortcuts import render,get_object_or_404
from django.utils.text import slugify

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from rest_framework_api_key.permissions import HasAPIKey

from api.authentication import SessionTokenAuthentication
from core.permissions import IsAdminAccess

from orgs.models import OrganizationMembership, Organization
from academics.models import (
    StudentProfile,
    Classroom,
    TeacherProfile,
    ParentProfile,
    ParentChildLink,
    Subject
)

from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password

from learning.models import (
    Module,
    Lesson,
    Course,
    Enrollment,
    CoursePassCriteria
)

from gamification.models import LeaderboardSeason
from core.models import StudentDevice, TimeStampedModel
from billing.models import OrganizationSubscription
from accounts.models import AdminAccess

from store.models import Category, Product, ProductImage
from store.serializers import (
    CategorySerializer,
    ProductAdminSerializer,
    ProductImageSerializer,
)

from .utils import (
    _get_student_for_user,
    _is_org_admin_or_teacher,
    _season_to_dict,
    _parse_dt,
    _resolve_org,
    _criteria_to_dict,
    _parse_positive_int,
    _get_user_avatar_url,
    _try_fetch_courses_for_classroom,
    _get_admin_selected_org_id
)

# academics/api.py
import csv
import io
import re
import secrets
from datetime import datetime
from decimal import Decimal

from django.utils import timezone
from django.http import HttpResponse


User = get_user_model()



def school_students_view(request, org_id):
    organization = get_object_or_404(
        Organization,
        pk=org_id
    )

    students = (
        StudentProfile.objects
        .select_related("user")
        .filter(organization=organization)
        .order_by("user__first_name", "user__last_name")
    )

    context = {
        "organization": organization,
        "students": students,
    }

    return render(
        request,
        "academics/school_students.html",
        context
    )

@api_view(["GET"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def active_modules_for_user(request):
    """
    Return all *active* Modules connected to the authenticated student.

    Scoping:
      - Only Modules with Module.active=True
      - Only Modules belonging to Courses the student is actively enrolled in
      - Course.is_active is also respected (optional, toggle with ?only_active_courses=0 to ignore)

    Query params:
      - q: text search (module name, course name, subject)
      - course_id: filter to a specific course
      - subject_id: filter to a subject
      - teacher_id: filter to a teacher profile id
      - page (default 1), page_size (default 20, max 100)
      - only_active_courses (default 1): if 1, restrict to Course.is_active=True
      - debug=1 to include traceback in error responses

    Response:
      {
        "count": <int>,
        "page": <int>,
        "page_size": <int>,
        "results": [
          {
            "id": <module_id>,
            "name": "...",
            "order": <int>,
            "active": true,
            "course": {
              "id": <int>, "name": "...", "subject": "...",
              "classroom": "...", "teacher": "Display Name"
            },
            "lessons_count": <int>,             # active lessons in the module
            "last_updated": "ISO8601 or null",   # last lesson update in the module
            "course_progress": <int>,            # % from Enrollment.progress_pct
            "recent_lesson": {                   # first lesson by order (active) if any
              "id": <int>,
              "title": "...",
              "content_type": "video|audio|pdf|doc|link",
              "duration_seconds": <int>,
              "url": "file/url/or/external/url"
            }
          },
          ...
        ]
      }
    """
    try:
        user = request.user
        student = _get_student_for_user(user)
        if not student:
            return Response(
                {"count": 0, "page": 1, "page_size": 0, "results": [], "detail": "Student profile not found."},
                status=status.HTTP_200_OK,
            )

        # ----------- parsing filters / pagination -----------
        def _i(v, d, cap=None):
            try:
                x = int(v) if v is not None else d
                return min(x, cap) if cap else x
            except Exception:
                return d

        q = (request.query_params.get("q") or "").strip()
        course_id = request.query_params.get("course_id")
        subject_id = request.query_params.get("subject_id")
        teacher_id = request.query_params.get("teacher_id")
        only_active_courses = _i(request.query_params.get("only_active_courses"), 1)

        page = _i(request.query_params.get("page"), 1)
        page_size = _i(request.query_params.get("page_size"), 20, cap=100)

        # ----------- student enrollments -----------
        enroll_qs = Enrollment.objects.filter(student=student, status__in=[Enrollment.Status.ACTIVE, Enrollment.Status.COMPLETED])
        if only_active_courses:
            enroll_qs = enroll_qs.filter(course__is_active=True)

        enrolled_course_ids = list(enroll_qs.values_list("course_id", flat=True))
        if not enrolled_course_ids:
            return Response({"count": 0, "page": page, "page_size": page_size, "results": []},
                            status=status.HTTP_200_OK)

        progress_map = {
            e.course_id: int(e.progress_pct or 0)
            for e in enroll_qs
        }

        # ----------- base modules queryset -----------
        modules_qs = (
            Module.objects
            .filter(active=True, course_id__in=enrolled_course_ids)
            .select_related("course", "course__subject", "course__classroom", "course__teacher__user")
            .annotate(
                lessons_count=Count("lessons", filter=Q(lessons__active=True), distinct=True),
                last_updated=Max("lessons__updated_at"),
            )
        )

        if course_id:
            modules_qs = modules_qs.filter(course_id=course_id)
        if subject_id:
            modules_qs = modules_qs.filter(course__subject_id=subject_id)
        if teacher_id:
            modules_qs = modules_qs.filter(course__teacher_id=teacher_id)
        if q:
            modules_qs = modules_qs.filter(
                Q(name__icontains=q) |
                Q(course__name__icontains=q) |
                Q(course__subject__name__icontains=q)
            )

        modules_qs = modules_qs.order_by("course_id", "order", "id")

        total = modules_qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        modules_page = list(modules_qs[start:end])

        # ----------- fetch a recent/first active lesson per module -----------
        module_ids = [m.id for m in modules_page]
        latest_by_module: Dict[int, Lesson] = {}
        if module_ids:
            # first active lesson by (order, id)
            for ls in (Lesson.objects
                       .filter(active=True, module_id__in=module_ids)
                       .order_by("module_id", "order", "id")):
                if ls.module_id not in latest_by_module:
                    latest_by_module[ls.module_id] = ls

        # ----------- build response -----------
        results: List[Dict[str, Any]] = []
        for m in modules_page:
            c: Course = m.course
            teacher_user = getattr(c.teacher, "user", None) if c and c.teacher else None
            teacher_name = teacher_user.get_full_name() if teacher_user else None
            subj_name = getattr(c.subject, "name", None) if c and c.subject else None
            classroom_name = getattr(c.classroom, "name", None) if c and c.classroom else None

            rl = latest_by_module.get(m.id)
            recent = None
            if rl:
                # link to file or external url
                url = None
                if rl.file:
                    try:
                        url = rl.file.url
                    except Exception:
                        url = None
                if not url:
                    url = rl.url or None

                recent = {
                    "id": rl.id,
                    "title": rl.name,
                    "content_type": rl.content_type,
                    "duration_seconds": int(rl.duration_seconds or 0),
                    "url": url,
                }

            results.append({
                "id": m.id,
                "name": m.name,
                "order": m.order,
                "active": m.active,
                "course": {
                    "id": c.id if c else None,
                    "name": c.name if c else None,
                    "subject": subj_name,
                    "classroom": classroom_name,
                    "teacher": teacher_name,
                },
                "lessons_count": int(getattr(m, "lessons_count", 0) or 0),
                "last_updated": m.last_updated.isoformat() if getattr(m, "last_updated", None) else None,
                "course_progress": progress_map.get(c.id if c else None, 0),
                "recent_lesson": recent,
            })

        return Response(
            {"count": total, "page": page, "page_size": page_size, "results": results},
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        err = {"detail": "Failed to load active modules for user.", "error": f"{type(e).__name__}: {e}"}
        if request.query_params.get("debug") in {"1", "true", "True"} or getattr(settings, "DEBUG", False):
            err["traceback"] = traceback.format_exc()
        return Response(err, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(["GET", "POST"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def leaderboard_seasons_view(request):
    """
    GET  /api/admin/settings/leaderboard-seasons
    POST /api/admin/settings/leaderboard-seasons

    Headers:
      Authorization: Api-Key <YOUR_API_KEY>
      X-Session-Token: <session_token>

    Body (POST):
      {
        "name": "2026 Academic Year",
        "slug": "2026-academic-year",   # optional; will slugify(name) if omitted
        "start_at": "2026-01-01T00:00:00Z",
        "end_at": "2026-12-31T23:59:59Z",
        "is_active": false              # optional
      }
    """
    try:
        org, err = _resolve_org(request)
        if err:
            return err

        if not _is_org_admin_or_teacher(request, org):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        if request.method == "GET":
            qs = LeaderboardSeason.objects.filter(organization=org).order_by("-start_at", "-id")
            return Response([_season_to_dict(s) for s in qs])

        # POST (create)
        data = request.data or {}

        name = (data.get("name") or "").strip()
        if not name:
            return Response({"detail": "name is required."}, status=status.HTTP_400_BAD_REQUEST)

        slug = (data.get("slug") or "").strip()
        if not slug:
            slug = slugify(name)[:128]

        start_at = _parse_dt(data.get("start_at"))
        end_at = _parse_dt(data.get("end_at"))
        if not start_at or not end_at:
            return Response(
                {"detail": "start_at and end_at must be valid ISO 8601 datetimes."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if start_at >= end_at:
            return Response({"detail": "start_at must be before end_at."}, status=status.HTTP_400_BAD_REQUEST)

        is_active = bool(data.get("is_active", False))

        with transaction.atomic():
            # if making active, deactivate others first
            if is_active:
                LeaderboardSeason.objects.filter(organization=org, is_active=True).update(is_active=False)

            s = LeaderboardSeason.objects.create(
                organization=org,
                name=name,
                slug=slug,
                start_at=start_at,
                end_at=end_at,
                is_active=is_active,
            )

        return Response(_season_to_dict(s), status=status.HTTP_201_CREATED)

    except Exception as e:
        traceback.print_exc()
        return Response(
            {"detail": "Unexpected error", "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def leaderboard_season_detail_view(request, season_id: int):
    """
    GET    /api/admin/settings/leaderboard-seasons/<season_id>
    PATCH  /api/admin/settings/leaderboard-seasons/<season_id>
    DELETE /api/admin/settings/leaderboard-seasons/<season_id>

    Body (PATCH): any of
      {
        "name": "...",
        "slug": "...",
        "start_at": "ISO",
        "end_at": "ISO",
        "is_active": true/false
      }
    """
    try:
        org, err = _resolve_org(request)
        if err:
            return err

        if not _is_org_admin_or_teacher(request, org):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        try:
            s = LeaderboardSeason.objects.get(id=season_id, organization=org)
        except LeaderboardSeason.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        if request.method == "GET":
            return Response(_season_to_dict(s))

        if request.method == "DELETE":
            s.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

        # PATCH
        data = request.data or {}

        name = data.get("name", None)
        slug = data.get("slug", None)
        start_at_val = data.get("start_at", None)
        end_at_val = data.get("end_at", None)
        is_active_val = data.get("is_active", None)

        if name is not None:
            name = (name or "").strip()
            if not name:
                return Response({"detail": "name cannot be empty."}, status=status.HTTP_400_BAD_REQUEST)
            s.name = name
            # if slug not explicitly provided, keep existing slug

        if slug is not None:
            slug = (slug or "").strip()
            if not slug:
                return Response({"detail": "slug cannot be empty."}, status=status.HTTP_400_BAD_REQUEST)
            s.slug = slug

        if start_at_val is not None:
            start_at = _parse_dt(start_at_val)
            if not start_at:
                return Response({"detail": "start_at must be valid ISO 8601 datetime."}, status=status.HTTP_400_BAD_REQUEST)
            s.start_at = start_at

        if end_at_val is not None:
            end_at = _parse_dt(end_at_val)
            if not end_at:
                return Response({"detail": "end_at must be valid ISO 8601 datetime."}, status=status.HTTP_400_BAD_REQUEST)
            s.end_at = end_at

        # validate date range if either changed
        if s.start_at and s.end_at and s.start_at >= s.end_at:
            return Response({"detail": "start_at must be before end_at."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            if is_active_val is not None:
                make_active = bool(is_active_val)
                if make_active:
                    LeaderboardSeason.objects.filter(organization=org, is_active=True).exclude(id=s.id).update(is_active=False)
                s.is_active = make_active

            s.save()

        return Response(_season_to_dict(s), status=status.HTTP_200_OK)

    except Exception as e:
        traceback.print_exc()
        return Response(
            {"detail": "Unexpected error", "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def leaderboard_season_set_active_view(request, season_id: int):
    """
    POST /api/admin/settings/leaderboard-seasons/<season_id>/set-active
    Sets this season as the ONLY active season for the org (atomic).
    """
    try:
        org, err = _resolve_org(request)
        if err:
            return err

        if not _is_org_admin_or_teacher(request, org):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        try:
            s = LeaderboardSeason.objects.get(id=season_id, organization=org)
        except LeaderboardSeason.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        with transaction.atomic():
            LeaderboardSeason.objects.filter(organization=org, is_active=True).exclude(id=s.id).update(is_active=False)
            s.is_active = True
            s.save()

        return Response(_season_to_dict(s), status=status.HTTP_200_OK)

    except Exception as e:
        traceback.print_exc()
        return Response(
            {"detail": "Unexpected error", "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )




@api_view(["GET", "POST", "PATCH"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def course_pass_criteria_view(request, course_id: int):
    """
    GET   /api/admin/courses/<course_id>/pass-criteria
    POST  /api/admin/courses/<course_id>/pass-criteria   (create or replace)
    PATCH /api/admin/courses/<course_id>/pass-criteria   (partial update)

    Headers:
      Authorization: Api-Key <YOUR_API_KEY>
      X-Session-Token: <session_token>

    Body (POST/PATCH):
      {
        "no_of_cbt": 10,
        "no_of_code_submission": 10,
        "total_pass_mark_cbt": 500,
        "total_pass_mark_code": 500
      }
    """
    try:
        org, err = _resolve_org(request)
        if err:
            return err

        if not _is_org_admin_or_teacher(request, org):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        course = get_object_or_404(Course, id=course_id, organization=org)

        # GET
        if request.method == "GET":
            criteria = getattr(course, "pass_criteria", None)
            if not criteria:
                # return defaults (optional) or 404. I prefer "empty + defaults" for UI simplicity.
                return Response(
                    {
                        "course_id": course.id,
                        "no_of_cbt": 10,
                        "no_of_code_submission": 10,
                        "total_pass_mark_cbt": 500,
                        "total_pass_mark_code": 500,
                        "exists": False,
                    },
                    status=status.HTTP_200_OK,
                )
            d = _criteria_to_dict(criteria)
            d["exists"] = True
            return Response(d, status=status.HTTP_200_OK)

        # POST/PATCH (upsert)
        data = request.data or {}

        # Defaults:
        default_no_of_cbt = 10
        default_no_of_code_submission = 10
        default_total_pass_mark_cbt = 500
        default_total_pass_mark_code = 500

        # For PATCH: only update provided fields
        is_patch = request.method == "PATCH"

        no_of_cbt, err_resp = _parse_positive_int(
            data.get("no_of_cbt"), "no_of_cbt", default=None if is_patch else default_no_of_cbt
        )
        if err_resp:
            return err_resp

        no_of_code_submission, err_resp = _parse_positive_int(
            data.get("no_of_code_submission"),
            "no_of_code_submission",
            default=None if is_patch else default_no_of_code_submission,
        )
        if err_resp:
            return err_resp

        total_pass_mark_cbt, err_resp = _parse_positive_int(
            data.get("total_pass_mark_cbt"),
            "total_pass_mark_cbt",
            default=None if is_patch else default_total_pass_mark_cbt,
        )
        if err_resp:
            return err_resp

        total_pass_mark_code, err_resp = _parse_positive_int(
            data.get("total_pass_mark_code"),
            "total_pass_mark_code",
            default=None if is_patch else default_total_pass_mark_code,
        )
        if err_resp:
            return err_resp

        with transaction.atomic():
            criteria, created = CoursePassCriteria.objects.get_or_create(course=course)

            # Apply fields (POST replaces all; PATCH only updates provided)
            if not is_patch or "no_of_cbt" in data:
                criteria.no_of_cbt = no_of_cbt if no_of_cbt is not None else criteria.no_of_cbt
            if not is_patch or "no_of_code_submission" in data:
                criteria.no_of_code_submission = (
                    no_of_code_submission
                    if no_of_code_submission is not None
                    else criteria.no_of_code_submission
                )
            if not is_patch or "total_pass_mark_cbt" in data:
                criteria.total_pass_mark_cbt = (
                    total_pass_mark_cbt
                    if total_pass_mark_cbt is not None
                    else criteria.total_pass_mark_cbt
                )
            if not is_patch or "total_pass_mark_code" in data:
                criteria.total_pass_mark_code = (
                    total_pass_mark_code
                    if total_pass_mark_code is not None
                    else criteria.total_pass_mark_code
                )

            criteria.save()

        out = _criteria_to_dict(criteria)
        out["created"] = created
        return Response(out, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

    except Exception as e:
        print(e)
        traceback.print_exc()
        return Response(
            {"detail": "Unexpected error", "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )




def _paginate(request, qs, serializer_cls, *, context=None):
    page = int(request.query_params.get("page", 1) or 1)
    page_size = int(request.query_params.get("page_size", 20) or 20)
    page_size = max(1, min(page_size, 100))

    total = qs.count()
    start = (page - 1) * page_size
    end = start + page_size
    rows = qs[start:end]

    ser = serializer_cls(rows, many=True, context=context or {})
    return Response({
        "count": total,
        "page": page,
        "page_size": page_size,
        "results": ser.data
    })


# -------------------------
# Categories CRUD
# -------------------------

@api_view(["GET", "POST"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_categories_list_create(request):
    if request.method == "GET":
        qs = Category.objects.all().order_by("name")
        return _paginate(request, qs, CategorySerializer, context={"request": request})

    ser = CategorySerializer(data=request.data, context={"request": request})
    ser.is_valid(raise_exception=True)
    obj = ser.save()
    return Response(CategorySerializer(obj, context={"request": request}).data, status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_categories_detail(request, category_id):
    obj = Category.objects.filter(id=category_id).first()
    if not obj:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(CategorySerializer(obj, context={"request": request}).data)

    if request.method == "PATCH":
        ser = CategorySerializer(obj, data=request.data, partial=True, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(CategorySerializer(obj, context={"request": request}).data)

    # DELETE
    obj.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


# -------------------------
# Products CRUD
# -------------------------

@api_view(["GET", "POST"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_products_list_create(request):
    if request.method == "GET":
        qs = Product.objects.select_related("category").prefetch_related("images").order_by("-created_at")

        q = (request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(title__icontains=q) |
                Q(slug__icontains=q) |
                Q(sku__icontains=q)
            )

        product_type = (request.query_params.get("product_type") or "").strip()
        if product_type:
            qs = qs.filter(product_type=product_type)

        category_id = (request.query_params.get("category_id") or "").strip()
        if category_id:
            qs = qs.filter(category_id=category_id)

        is_active = request.query_params.get("is_active")
        if is_active in ("true", "false"):
            qs = qs.filter(is_active=(is_active == "true"))

        return _paginate(request, qs, ProductAdminSerializer, context={"request": request})

    # POST create
    ser = ProductAdminSerializer(data=request.data, context={"request": request})
    ser.is_valid(raise_exception=True)
    obj = ser.save()
    return Response(ProductAdminSerializer(obj, context={"request": request}).data, status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_products_detail(request, product_id):
    obj = Product.objects.select_related("category").prefetch_related("images").filter(id=product_id).first()
    if not obj:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(ProductAdminSerializer(obj, context={"request": request}).data)

    if request.method == "PATCH":
        ser = ProductAdminSerializer(obj, data=request.data, partial=True, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(ProductAdminSerializer(obj, context={"request": request}).data)

    # DELETE
    obj.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


# -------------------------
# Product Images (upload + delete + reorder)
# -------------------------

@api_view(["POST"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
@transaction.atomic
def admin_product_images_upload(request, product_id):
    product = Product.objects.filter(id=product_id).first()
    if not product:
        return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)

    # multipart expected: file in "product_image"
    file = request.FILES.get("product_image")
    if not file:
        return Response({"detail": "product_image file is required."}, status=status.HTTP_400_BAD_REQUEST)

    alt_text = request.data.get("alt_text", "") or ""
    sort_order = int(request.data.get("sort_order") or 0)

    img = ProductImage.objects.create(
        product=product,
        product_image=file,
        alt_text=alt_text,
        sort_order=sort_order,
    )

    return Response(ProductImageSerializer(img, context={"request": request}).data, status=status.HTTP_201_CREATED)


@api_view(["DELETE", "PATCH"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_product_images_detail(request, image_id):
    img = ProductImage.objects.select_related("product").filter(id=image_id).first()
    if not img:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "PATCH":
        # allow alt_text / sort_order edits
        ser = ProductImageSerializer(img, data=request.data, partial=True, context={"request": request})
        ser.is_valid(raise_exception=True)
        img = ser.save()
        return Response(ProductImageSerializer(img, context={"request": request}).data)

    img.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)




@api_view(["GET"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_classroom_modal_data(request, classroom_id: int):
    try:
        """
        GET /api/admin/classrooms/<id>/modal/

        Returns ONLY what the ClassroomDetailsModal needs.
        ✅ No Student Grade included.
        """
        classroom = (
            Classroom.objects.select_related("organization")
            .prefetch_related("teachers")
            .filter(id=classroom_id)
            .first()
        )
        if not classroom:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Students in this classroom
        students_qs = (
            StudentProfile.objects.select_related("user")
            .filter(organization_id=classroom.organization_id, current_classroom_id=classroom.id)
            .order_by("user__first_name", "user__last_name")
        )

        students = [
            {
                "id": s.id,
                "user_id": s.user_id,
                "name": (s.user.get_full_name() or s.user.email or ""),
                "email": getattr(s.user, "email", "") or "",
                "avatar_url": _get_user_avatar_url(request, s.user),
                "admission_no": s.admission_no or "",
            }
            for s in students_qs
        ]

        # Teachers attached to classroom (many-to-many)
        teacher_user_ids = list(classroom.teachers.values_list("id", flat=True))
        teachers_profiles = (
            TeacherProfile.objects.select_related("user")
            .prefetch_related("specialties")
            .filter(organization_id=classroom.organization_id, user_id__in=teacher_user_ids)
        )
        # map for quick lookup
        tp_by_user = {tp.user_id: tp for tp in teachers_profiles}

        teachers = []
        for u in classroom.teachers.all():
            tp = tp_by_user.get(u.id)
            teachers.append(
                {
                    "user_id": u.id,
                    "name": u.get_full_name(),
                    "email": getattr(u, "email", "") or "",
                    "avatar_url": _get_user_avatar_url(request, u),
                    "specialties": [s.name for s in (tp.specialties.all() if tp else [])],
                }
            )

        # Courses (optional inference)
        courses, courses_count = _try_fetch_courses_for_classroom(classroom)

        payload = {
            "id": classroom.id,
            "name": classroom.name,
            "code": classroom.code,
            "description": getattr(classroom, "description", "") or "",
            "stats": {
                "students": len(students),
                "teachers": len(teachers),
                "courses": courses_count,
            },
            "students": students,  # ✅ no grade field here
            "teachers": teachers,
            "courses": courses,
        }
        return Response(payload, status=status.HTTP_200_OK)

    except Exception as e:
        print(e)
    return Response({}, status=status.HTTP_400_BAD_REQUEST)




def _device_dict(d: StudentDevice):
    return {
        "id": d.id,
        "device_id": d.device_id,
        "user_agent": d.user_agent,
        "first_seen": d.first_seen,
        "last_seen": d.last_seen,
    }


def _student_result(sp: StudentProfile):
    u = sp.user
    full_name = (u.get_full_name() or "").strip()
    return {
        "student_id": sp.id,
        "user_id": u.id,
        "email": u.email,
        "full_name": full_name or None,
        "organization_id": sp.organization_id,
        "devices": [_device_dict(d) for d in sp.devices.order_by("-last_seen")],
    }


@api_view(["GET"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_student_devices_search(request):
    """
    GET /api/admin/student-devices?query=...&limit=20
    Search by student email or name, return students + their devices.
    Scoped to adminaccess.selected_organization.
    """
    org_id = _get_admin_selected_org_id(request)
    if not org_id:
        return Response(
            {"detail": "No selected organization for this admin."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    query = (request.query_params.get("query") or "").strip()
    limit = int(request.query_params.get("limit") or 20)
    limit = max(1, min(limit, 50))

    if not query:
        return Response({"count": 0, "results": []}, status=status.HTTP_200_OK)

    # Search students in selected org by email or name
    qs = (
        StudentProfile.objects
        .select_related("user")
        .prefetch_related("devices")
        .filter(organization_id=org_id)
        .filter(
            Q(user__email__icontains=query)
            | Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
            | Q(user__first_name__icontains=query.split(" ")[0])  # helps partials
        )
        .order_by("user__first_name", "user__last_name")
    )

    students = list(qs[:limit])
    return Response(
        {"count": qs.count(), "results": [_student_result(sp) for sp in students]},
        status=status.HTTP_200_OK,
    )


@api_view(["DELETE"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
@transaction.atomic
def admin_student_device_delete(request, device_pk: int):
    """
    DELETE /api/admin/student-devices/<device_pk>/
    Deletes one StudentDevice (scoped to admin selected org).
    """
    org_id = _get_admin_selected_org_id(request)
    if not org_id:
        return Response(
            {"detail": "No selected organization for this admin."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    device = (
        StudentDevice.objects
        .select_related("student", "student__user")
        .filter(id=device_pk, student__organization_id=org_id)
        .first()
    )
    if not device:
        return Response({"detail": "Device not found."}, status=status.HTTP_404_NOT_FOUND)

    device.delete()
    return Response({"detail": "Deleted."}, status=status.HTTP_200_OK)




TEMPLATE_HEADERS = [
    "parent_first",
    "parent_last",
    "parent_email",
    "parent_phone",
    "parent_address",
    "student_first",
    "student_last",
    "student_email",
    "student_dob(YYYY-MM-DD)",
    "relationship",
    "teacher_email",
    "course_name",
    "classroom_name",
    "subject_name",
    "class_type",  # public|private
]


def _make_generated_email(first: str, last: str, domain: str = "testtechxagonacademy.com"):
    """Generate deterministic-looking email from names; remove special chars."""
    base = f"{(first or '')}{(last or '')}".lower()
    # remove non-alphanumeric
    base = re.sub(r"[^a-z0-9]", "", base)
    if not base:
        base = "user" + secrets.token_hex(3)
    email = f"{base}@{domain}"
    # ensure unique by appending numeric suffix if needed
    attempt = 0
    candidate = email
    while User.objects.filter(email=candidate).exists():
        attempt += 1
        candidate = f"{base}{attempt}@{domain}"
    return candidate


def _get_request_org(request):
    """Return the Organization of the current logged-in user (first active membership)."""
    try:
        membership = OrganizationMembership.objects.filter(user=request.user, is_active=True).first()
        if membership:
            return membership.organization
    except Exception:
        pass
    # fallback: if request.user has organization relation etc
    org = getattr(request.user, "organization", None)
    return org


@api_view(["GET"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])  # adjust per your project, you had SessionTokenAuthentication
def download_csv_template(request):
    """Return a CSV template for admins to download and fill."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(TEMPLATE_HEADERS)
    # Provide an example row (optional)
    writer.writerow([
        "Jane", "Doe", "jane.doe@example.com", "+2348012345678", "12 Main St, Ikeja",
        "John", "Doe", "", "2014-05-01", "mother", "teacher@example.com", "Intro to Math", "Primary 1", "Mathematics", "public"
    ])

    resp = HttpResponse(output.getvalue(), content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="parent_student_template.csv"'
    return resp


@api_view(["POST"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication]) 
def upload_parent_student_csv(request):
    """
    Accepts a multipart/form-data file under key "file" or raw CSV body.
    CSV columns (header expected):
    parent_first,parent_last,parent_email,parent_phone,parent_address,
    student_first,student_last,student_email,student_dob(YYYY-MM-DD),relationship,
    teacher_email,course_name,classroom_name,subject_name,class_type
    """
    file_obj = request.FILES.get("file") or request.data.get("file")
    if not file_obj:
        return Response({"error": "CSV file is required in 'file' field."}, status=status.HTTP_400_BAD_REQUEST)

    # Determine organization from request user
    org = _get_request_org(request)
    if org is None:
        return Response({"error": "Could not determine organization for the current user."}, status=status.HTTP_400_BAD_REQUEST)

    # --- NEW: require active leaderboard season BEFORE doing any creation ---
    active_season = LeaderboardSeason.get_active(org=org)
    if active_season is None:
        return Response(
            {"error": "No active leaderboard season found for this organisation. Aborting import. Please activate a LeaderboardSeason before importing."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    decoded = file_obj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded))

    header = reader.fieldnames or []
    missing = [h for h in TEMPLATE_HEADERS if h not in header]
    if missing:
        return Response({"error": "Missing columns", "missing": missing}, status=status.HTTP_400_BAD_REQUEST)

    # Caches to avoid repeated DB lookups
    parents_cache = {}  # key by email or name+phone -> ParentProfile
    students_cache = {}
    teachers_cache = {}
    classrooms_cache = {}
    subjects_cache = {}
    courses_created = {}
    created_users = []
    created_parents = []
    created_students = []
    created_teachers = []
    created_courses = []
    created_enrollments = []
    errors = []
    row_no = 1

    with transaction.atomic():
        for row in reader:
            row_no += 1
            try:
                p_first = (row.get("parent_first") or "").strip()
                p_last = (row.get("parent_last") or "").strip()
                p_email = (row.get("parent_email") or "").strip()
                p_phone = (row.get("parent_phone") or "").strip()
                p_address = (row.get("parent_address") or "").strip()

                s_first = (row.get("student_first") or "").strip()
                s_last = (row.get("student_last") or "").strip()
                s_email = (row.get("student_email") or "").strip()
                s_dob = (row.get("student_dob(YYYY-MM-DD)") or "").strip()
                relationship = (row.get("relationship") or "").strip() or "parent"

                teacher_email = (row.get("teacher_email") or "").strip()
                course_name = (row.get("course_name") or "").strip()
                classroom_name = (row.get("classroom_name") or "").strip()
                subject_name = (row.get("subject_name") or "").strip()
                class_type = (row.get("class_type") or "public").strip()

                if not (p_first or p_last):
                    raise ValueError("Parent name missing")
                if not (s_first or s_last):
                    raise ValueError("Student name missing")

                # Parent
                parent_key = p_email.lower() if p_email else f"{p_first.lower()}_{p_last.lower()}_{p_phone}"
                parent_profile = parents_cache.get(parent_key)
                if not parent_profile:
                    parent_user = None
                    if p_email:
                        parent_user = User.objects.filter(email__iexact=p_email).first()
                    if not parent_user:
                        generated_email = p_email or _make_generated_email(p_first, p_last)
                        base_email = generated_email
                        attempt = 0
                        while User.objects.filter(email=generated_email).exists():
                            attempt += 1
                            generated_email = f"{slugify(p_first + p_last or 'user') or 'user'}{attempt}@testtechxagonacademy.com"
                        parent_user = User.objects.create(
                            email=generated_email,
                            first_name=p_first,
                            last_name=p_last,
                            password=make_password(secrets.token_hex(8)),
                        )
                        created_users.append(parent_user)
                        try:
                            setattr(parent_user, "is_generated", True)
                            parent_user.save(update_fields=["is_generated"])
                        except Exception:
                            parent_user.save()

                    parent_profile_obj, _ = ParentProfile.objects.get_or_create(
                        user=parent_user,
                        defaults={"organization": org, "address": p_address}
                    )
                    if p_address and parent_profile_obj.address != p_address:
                        parent_profile_obj.address = p_address
                        parent_profile_obj.save(update_fields=["address"])

                    OrganizationMembership.objects.get_or_create(
                        user=parent_user,
                        organization=org,
                        role=OrganizationMembership.Role.PARENT,
                        defaults={"is_active": True},
                    )
                    parents_cache[parent_key] = parent_profile_obj
                    parent_profile = parent_profile_obj
                    created_parents.append(parent_profile)

                # Student
                student_key = s_email.lower() if s_email else f"{s_first.lower()}_{s_last.lower()}_{parent_key}_{s_dob}"
                student_profile = students_cache.get(student_key)
                if not student_profile:
                    student_user = None
                    if s_email:
                        student_user = User.objects.filter(email__iexact=s_email).first()
                    if not student_user:
                        gen_email = s_email or _make_generated_email(s_first, s_last)
                        attempt = 0
                        candidate = gen_email
                        while User.objects.filter(email=candidate).exists():
                            attempt += 1
                            candidate = f"{slugify(s_first + s_last or 'student') or 'student'}{attempt}@testtechxagonacademy.com"
                        student_user = User.objects.create(
                            email=candidate,
                            first_name=s_first,
                            last_name=s_last,
                            password=make_password(secrets.token_hex(8)),
                        )
                        created_users.append(student_user)
                        try:
                            setattr(student_user, "is_generated", True)
                            student_user.save(update_fields=["is_generated"])
                        except Exception:
                            student_user.save()

                    stu_defaults = {"organization": org}
                    if s_dob:
                        try:
                            stu_defaults["dob"] = datetime.strptime(s_dob, "%Y-%m-%d").date()
                        except Exception:
                            pass

                    student_profile_obj, _ = StudentProfile.objects.get_or_create(
                        user=student_user,
                        defaults=stu_defaults
                    )
                    if s_dob and (not student_profile_obj.dob):
                        try:
                            student_profile_obj.dob = datetime.strptime(s_dob, "%Y-%m-%d").date()
                            student_profile_obj.save(update_fields=["dob"])
                        except Exception:
                            pass

                    OrganizationMembership.objects.get_or_create(
                        user=student_user,
                        organization=org,
                        role=OrganizationMembership.Role.STUDENT,
                        defaults={"is_active": True},
                    )

                    students_cache[student_key] = student_profile_obj
                    student_profile = student_profile_obj
                    created_students.append(student_profile)

                ParentChildLink.objects.get_or_create(
                    parent=parent_profile,
                    student=student_profile,
                    defaults={"relationship": relationship}
                )

                # Teacher
                teacher = None
                if teacher_email:
                    tkey = teacher_email.lower()
                    teacher = teachers_cache.get(tkey)
                    if not teacher:
                        t_user = User.objects.filter(email__iexact=teacher_email).first()
                        if not t_user:
                            t_user = User.objects.create(
                                email=teacher_email,
                                first_name="",
                                last_name="",
                                password=make_password(secrets.token_hex(8)),
                            )
                            created_users.append(t_user)
                            try:
                                setattr(t_user, "is_generated", True)
                                t_user.save(update_fields=["is_generated"])
                            except Exception:
                                t_user.save()

                        teacher_profile_obj, _ = TeacherProfile.objects.get_or_create(
                            user=t_user,
                            organization=org,
                            defaults={"bio": ""}
                        )
                        OrganizationMembership.objects.get_or_create(
                            user=t_user,
                            organization=org,
                            role=OrganizationMembership.Role.TEACHER,
                            defaults={"is_active": True},
                        )
                        teachers_cache[tkey] = teacher_profile_obj
                        teacher = teacher_profile_obj

                # Classroom
                classroom = None
                if classroom_name:
                    ckey = f"{org.id}::{classroom_name}::{class_type}"
                    classroom = classrooms_cache.get(ckey)
                    if not classroom:
                        classroom_obj, _ = Classroom.objects.get_or_create(
                            organization=org,
                            name=classroom_name,
                            defaults={"code": slugify(classroom_name)[:32], "class_type": class_type or "public"}
                        )
                        classrooms_cache[ckey] = classroom_obj
                        classroom = classroom_obj

                # Subject
                subject = None
                if subject_name:
                    skey = f"{org.id}::{subject_name}"
                    subject = subjects_cache.get(skey)
                    if not subject:
                        subject_obj, _ = Subject.objects.get_or_create(
                            organization=org,
                            name=subject_name,
                            defaults={"code": slugify(subject_name)[:32]}
                        )
                        subjects_cache[skey] = subject_obj
                        subject = subject_obj

                # Course
                course_obj = None
                if course_name:
                    # If teacher/classroom/subject present, try to find or create
                    # If any dependency is missing we still attempt creation but prefer full set
                    t_id = getattr(teacher, "id", None)
                    c_id = getattr(classroom, "id", None)
                    s_id = getattr(subject, "id", None)

                    ckey = f"{org.id}::{course_name}::{s_id}::{c_id}::{t_id}"
                    if ckey not in courses_created:
                        course_defaults = {
                            "name": course_name,
                            "description": f"Auto-created course {course_name}",
                            "is_active": True,
                            "course_type": class_type or "public",
                        }
                        # Try to set foreign keys only if available
                        course_filter = {"organization": org}
                        if s_id:
                            course_filter["subject_id"] = s_id
                        if c_id:
                            course_filter["classroom_id"] = c_id
                        if t_id:
                            course_filter["teacher_id"] = t_id

                        course_obj, created_flag = Course.objects.get_or_create(
                            defaults=course_defaults,
                            **course_filter
                        ) if course_filter else Course.objects.get_or_create(
                            organization=org,
                            subject=subject,
                            classroom=classroom,
                            teacher=teacher,
                            defaults=course_defaults
                        )
                        # above: attempt to be flexible to find existing; adjust to your project's preferred lookup
                        courses_created[ckey] = course_obj
                        if created_flag:
                            created_courses.append(course_obj)
                    else:
                        course_obj = courses_created[ckey]

                # --- ENROLLMENT: if course exists (created or previously existed), create Enrollment for active season ---
                if course_obj:
                    enrollment_obj, created_enrolled = Enrollment.objects.get_or_create(
                        student=student_profile,
                        course=course_obj,
                        defaults={
                            "leaderboard_season": active_season,
                            "status": Enrollment.Status.ACTIVE,
                            "progress_pct": 0,
                        },
                    )
                    # If enrollment existed but had null leaderboard_season, try to set it (only if null)
                    if not created_enrolled:
                        if enrollment_obj.leaderboard_season is None:
                            enrollment_obj.leaderboard_season = active_season
                            enrollment_obj.save(update_fields=["leaderboard_season"])
                    if created_enrolled:
                        created_enrollments.append(enrollment_obj)

            except Exception as e:
                errors.append({"row": row_no, "error": str(e), "row": row})
                # continue to next row

    summary = {
        "rows_processed": row_no - 1,
        "users_created": len(created_users),
        "parents_created": len(created_parents),
        "students_created": len(created_students),
        "teachers_created": len(created_teachers),
        "courses_created": len(created_courses),
        "enrollments_created": len(created_enrollments),
        "errors": errors,
    }
    status_code = status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED
    return Response(summary, status=status_code)


# ── Login Generation ─────────────────────────────────────────────────────────

import unicodedata
from datetime import timedelta
from openpyxl import Workbook, load_workbook
from io import BytesIO
from billing.models import SubscriptionPlan, UserAccountSubscription


def _sanitise_for_email(full_name: str) -> str:
    """
    Convert a full name to a lowercase, ASCII-only, punctuation-free string
    suitable for use as an email local-part.
    """
    nfkd = unicodedata.normalize("NFKD", full_name)
    ascii_only = nfkd.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", ascii_only.lower())


@api_view(["GET"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def login_generation_template(request):
    """
    GET /core/api/admin/login-generation/template/
    Returns an .xlsx template with columns: Student Name
    Each sheet tab represents a classroom.
    """
    try:
        org, err = _resolve_org(request)
        if err:
            return err

        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["Student Full Name"])
        # Add a couple of example rows
        ws.append(["John Doe"])
        ws.append(["Jane Smith"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student_login_template.xlsx"'
        return response

    except Exception as e:
        traceback.print_exc()
        return Response(
            {"detail": "Failed to generate template.", "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([HasAPIKey])
@authentication_classes([SessionTokenAuthentication])
def generate_student_logins(request):
    """
    POST /core/api/admin/login-generation/generate/

    Accepts two modes:
      1) mode=excel  → multipart with file= .xlsx
      2) mode=manual → JSON body with students=[{name, classroom_name}]

    Required: classroom_id (for manual mode per-student or a single one)

    Creates for each student:
      • User (email = <sanitised-name><org-slug>@techxagonacademy.com, password = Techxagon@2026)
      • OrganizationMembership (role=student)
      • StudentProfile (org + classroom linked)
      • UserAccountSubscription (Free plan, 1 month, active)

    Everything runs inside transaction.atomic() — any error rolls back all.

    Returns JSON with created student details (name, email, password, classroom, admission_no).
    """
    try:
        org, err = _resolve_org(request)
        if err:
            return err

        if not _is_org_admin_or_teacher(request, org):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        mode = request.data.get("mode") or request.POST.get("mode") or "manual"

        # ── Resolve the Free plan (SubscriptionPlan with price=0 or lowest price)
        free_plan = SubscriptionPlan.objects.filter(price=0).first()
        if not free_plan:
            free_plan = SubscriptionPlan.objects.order_by("price").first()
        if not free_plan:
            return Response(
                {"detail": "No subscription plan found. Please create a Free plan first."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        password = "Techxagon@2026"
        now = timezone.now()
        end_at = now + timedelta(days=30)  # 1 month subscription
        # Build a short org tag for the email local-part.
        # Multi-word names get abbreviated (first letter of each word).
        # E.g. "Anglican Comprehensive Secondary School Kubwa" → "acssk"
        raw_org_name = (org.name or org.slug or "org")
        org_words = raw_org_name.split()
        if len(org_words) > 1:
            org_slug = "".join(w[0] for w in org_words if w).lower()
        else:
            org_slug = re.sub(r"[^a-z0-9]", "", raw_org_name.lower())


        # ── Parse students list based on mode ─────────────────────────
        students_to_create = []  # list of dicts: {name, classroom_name}

        if mode == "excel":
            file = request.FILES.get("file")
            if not file:
                return Response(
                    {"detail": "Excel file is required for mode=excel."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:
                wb = load_workbook(file, read_only=True)
            except Exception as exc:
                return Response(
                    {"detail": f"Cannot open Excel file: {exc}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            classroom_id = request.data.get("classroom_id") or request.POST.get("classroom_id")
            if not classroom_id:
                return Response(
                    {"detail": "classroom_id is required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:
                classroom_obj = Classroom.objects.get(pk=int(classroom_id), organization=org)
            except (Classroom.DoesNotExist, ValueError):
                return Response(
                    {"detail": f"Classroom with id={classroom_id} not found in this organization."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                for row in rows[1:]:  # skip header row
                    if not row or len(row) < 1:
                        continue
                    name = str(row[0] or "").strip()
                    if not name:
                        continue
                    students_to_create.append({
                        "name": name,
                        "classroom_name": classroom_obj.name,
                        "classroom_id": classroom_obj.id,
                    })

            wb.close()

        elif mode == "manual":
            students_raw = request.data.get("students", [])
            if not students_raw or not isinstance(students_raw, list):
                return Response(
                    {"detail": "students list is required for mode=manual."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for s in students_raw:
                name = (s.get("name") or "").strip()
                classroom_id = s.get("classroom_id")
                if not name:
                    continue
                if not classroom_id:
                    return Response(
                        {"detail": f"classroom_id is required for student '{name}'."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                try:
                    cls_obj = Classroom.objects.get(pk=int(classroom_id), organization=org)
                except (Classroom.DoesNotExist, ValueError):
                    return Response(
                        {"detail": f"Classroom id={classroom_id} not found for student '{name}'."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                students_to_create.append({
                    "name": name,
                    "classroom_name": cls_obj.name,
                    "classroom_id": cls_obj.id,
                })
        else:
            return Response(
                {"detail": "Invalid mode. Use 'excel' or 'manual'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not students_to_create:
            return Response(
                {"detail": "No valid students found to process."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Seed email uniqueness pool ────────────────────────────
        existing_emails = set(
            User.objects.filter(email__endswith="@techxagonacademy.com")
            .values_list("email", flat=True)
        )
        assigned_emails = set()   # emails assigned in this batch

        # Track (sanitised_name, classroom_id) pairs to detect same-class duplicates
        seen_name_in_class: set[tuple[str, int]] = set()

        # ── Create everything atomically ──────────────────────────
        created_results = []
        duplicate_results = []   # same name in the same classroom
        stats = {
            "users_created": 0,
            "profiles_created": 0,
            "memberships_created": 0,
            "subscriptions_created": 0,
            "skipped_existing": 0,
            "duplicates": 0,
        }

        # Pre-fetch / cache classrooms
        classroom_cache = {}

        with transaction.atomic():
            for student_data in students_to_create:
                full_name = student_data["name"]
                cls_id = student_data["classroom_id"]
                cls_name = student_data["classroom_name"]

                # Get or cache classroom
                if cls_id not in classroom_cache:
                    classroom_cache[cls_id] = Classroom.objects.get(pk=cls_id, organization=org)
                classroom = classroom_cache[cls_id]

                # ── Build unique email ────────────────────
                base_local = _sanitise_for_email(full_name)
                if not base_local:
                    base_local = "student"

                # ── Same-classroom duplicate check ────────
                # If we have already seen this exact name for this classroom
                # in this batch, it is a true duplicate — skip and record it.
                name_class_key = (base_local, cls_id)
                if name_class_key in seen_name_in_class:
                    stats["duplicates"] += 1
                    duplicate_results.append({
                        "name": full_name,
                        "classroom": cls_name,
                        "reason": "Duplicate name in the same classroom — skipped.",
                    })
                    continue
                seen_name_in_class.add(name_class_key)

                # ── Cross-classroom collision: auto-suffix ──
                local_part = base_local + org_slug
                base_email = f"{local_part}@techxagonacademy.com"

                # KEY FIX: if the base email (no suffix) already exists AND
                # that user is already an active member of this org, this is a
                # re-submission of an existing student — skip, don't mint a
                # new suffixed account (aguchimdikeacssk1, acssk2, etc.).
                if base_email in existing_emails:
                    already_in_org = OrganizationMembership.objects.filter(
                        user__email=base_email,
                        organization=org,
                        is_active=True,
                    ).exists()
                    if already_in_org:
                        stats["skipped_existing"] += 1
                        continue
                    # base email exists but NOT in this org → genuine name
                    # collision (different org), fall through to suffix loop

                candidate = base_email
                suffix = 0
                while candidate in existing_emails or candidate in assigned_emails:
                    suffix += 1
                    candidate = f"{local_part}{suffix}@techxagonacademy.com"

                email = candidate
                assigned_emails.add(email)
                existing_emails.add(email)

                # ── Safety: skip if somehow the final candidate is in DB ──
                if User.objects.filter(email=email).exists():
                    stats["skipped_existing"] += 1
                    continue


                # ── Split name ────────────────────────────
                parts = full_name.strip().split()
                first_name = parts[0] if parts else ""
                last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

                # ── User ──────────────────────────────────
                user = User.objects.create_user(
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_generated=True,
                    is_active=True,
                    primary_org=org,
                )
                stats["users_created"] += 1

                # ── OrganizationMembership ────────────────
                OrganizationMembership.objects.create(
                    user=user,
                    organization=org,
                    role=OrganizationMembership.Role.STUDENT,
                    is_active=True,
                )
                stats["memberships_created"] += 1

                # ── StudentProfile ────────────────────────
                sp = StudentProfile.objects.create(
                    user=user,
                    organization=org,
                    current_classroom=classroom,
                )
                stats["profiles_created"] += 1

                # ── UserAccountSubscription (1 month Free) ─
                UserAccountSubscription.objects.create(
                    organization=org,
                    user=user,
                    plan=free_plan,
                    status=UserAccountSubscription.Status.ACTIVE,
                    start_at=now,
                    end_at=end_at,
                    amount=free_plan.price,
                    currency="NGN",
                )
                stats["subscriptions_created"] += 1

                created_results.append({
                    "name": full_name,
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                    "password": password,
                    "classroom": cls_name,
                    "admission_no": sp.admission_no or "",
                })

        return Response(
            {
                "detail": "Login generation completed successfully.",
                "stats": stats,
                "students": created_results,
                "duplicates": duplicate_results,
            },
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        traceback.print_exc()
        return Response(
            {"detail": "Login generation failed. All changes have been rolled back.", "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# ── Admin: Password Management ─────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_list_org_users(request):
    """
    GET /core/api/admin/change-password/users/
    Returns all active members of the admin's selected organisation.
    Supports ?q= search by name or email.
    """
    org_id = _get_admin_selected_org_id(request)
    if not org_id:
        return Response(
            {"detail": "No selected organisation for this admin."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    q_param = (request.query_params.get("q") or "").strip()

    qs = (
        OrganizationMembership.objects
        .filter(organization_id=org_id, is_active=True)
        .select_related("user")
        .order_by("user__first_name", "user__last_name")
    )

    if q_param:
        qs = qs.filter(
            Q(user__first_name__icontains=q_param)
            | Q(user__last_name__icontains=q_param)
            | Q(user__email__icontains=q_param)
        )

    seen_ids: set[int] = set()
    users = []
    for m in qs:
        if m.user_id in seen_ids:
            continue
        seen_ids.add(m.user_id)
        users.append({
            "id": m.user_id,
            "name": (m.user.get_full_name() or m.user.email or ""),
            "email": m.user.email,
            "role": m.role,
            "avatar_url": _get_user_avatar_url(request, m.user),
        })

    return Response({"count": len(users), "results": users}, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([HasAPIKey, IsAdminAccess])
@authentication_classes([SessionTokenAuthentication])
def admin_change_user_password(request):
    """
    POST /core/api/admin/change-password/
    Body: { user_id: <int>, new_password: <str> }

    Changes the password for `user_id` only if that user belongs to the
    admin's selected organisation.
    """
    org_id = _get_admin_selected_org_id(request)
    if not org_id:
        return Response(
            {"detail": "No selected organisation for this admin."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user_id = request.data.get("user_id")
    new_password = (request.data.get("new_password") or "").strip()

    if not user_id:
        return Response({"detail": "user_id is required."}, status=status.HTTP_400_BAD_REQUEST)
    if not new_password or len(new_password) < 6:
        return Response(
            {"detail": "new_password must be at least 6 characters."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Verify the target user belongs to this org
    membership = OrganizationMembership.objects.filter(
        user_id=user_id,
        organization_id=org_id,
        is_active=True,
    ).select_related("user").first()

    if not membership:
        return Response(
            {"detail": "User not found in this organisation."},
            status=status.HTTP_404_NOT_FOUND,
        )

    target_user = membership.user
    target_user.set_password(new_password)
    target_user.save(update_fields=["password"])

    return Response(
        {
            "detail": f"Password updated for {target_user.get_full_name() or target_user.email}.",
            "user_id": target_user.id,
            "email": target_user.email,
        },
        status=status.HTTP_200_OK,
    )
